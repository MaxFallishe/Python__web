import telebot
from telebot import types
from openpyxl import load_workbook
import threading
import time


import socks,socket
# для socks4\5
socks.setdefaultproxy(socks.PROXY_TYPE_SOCKS5, '165.22.7.138', 8080)
# или для http\https
socks.setdefaultproxy(socks.PROXY_TYPE_HTTP, '165.22.7.138', 8080)
socket.socket = socks.socksocket



bot = telebot.TeleBot("xxx")

excel_row = 11
excel_row_2 = 2
wb = load_workbook('TicketsDataBase.xlsx')
ws = wb.active


@bot.message_handler(commands=['start'])
def start(message):
    bot.send_message(message.chat.id, "Вас приветствует TicketBot ")
    main(message)


@bot.message_handler(content_types=["text"])
def handle_text(message):
    global excel_row_2
    print("Отправил своё имя", message.chat.id, message.text)
    ws.cell(row=excel_row_2, column=8).value = message.chat.id
    ws.cell(row=excel_row_2, column=9).value = message.text
    wb.save("TicketsDataBase.xlsx")
    excel_row_2 += 1


def main(message):
    get_contacts(message)


def get_contacts(message):
    keyboard = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
    button_geo = types.KeyboardButton(text="Отправить номер", request_contact=True)
    keyboard.add(button_geo)
    bot.send_message(message.chat.id,
                     "Чтобы заказать билеты, сначала отправьте нам ваш номер телефона нажав на кнопку 'Отправить номер'",
                     reply_markup=keyboard)
    bot.register_next_step_handler(message, get_people_count)


def get_people_count(message):
    try:
        print(message.contact.phone_number)
        contact = message.contact.phone_number
        bot.send_message(message.chat.id, "Отправьте нам количество мест для бронирования",
                         reply_markup=types.ReplyKeyboardRemove())
        bot.register_next_step_handler(message, get_company_name, contact)
    except:
        print(1)
        bot.send_message(message.chat.id, "Пожалуйста нажмите на кнопку")
        get_contacts(message)


def get_company_name(message, contact):
    print(message.text)
    people_count = message.text
    bot.send_message(message.chat.id, "Отправьте название вашей компании")
    bot.register_next_step_handler(message, get_question, contact, people_count)


def get_question(message, contact, people_count):
    print(message.text)
    company_name = message.text
    bot.send_message(message.chat.id, "Отправьте вопрос который вы хотели бы задать, если вопроса нет, то напишиите 'Нет вопроса'")
    bot.register_next_step_handler(message, get_name, contact, people_count, company_name)


def get_name(message, contact, people_count, company_name):
    print(message.text)
    question = message.text
    bot.send_message(message.chat.id, "Отправьте ваше имя, или как мы можем к вам обращаться")
    bot.register_next_step_handler(message, write_in_excel, contact, people_count, company_name, question)


def write_in_excel(message, contact, people_count, company_name, question):
    global excel_row

    print(message.text)
    partner_name = message.text

    bot.send_message(message.chat.id, "Вы прошли регистрацию, ждём Вас 14-го июня, в 19:00 в GroundZero Chilanzar")
    ws.cell(row=excel_row, column=1).value = contact
    ws.cell(row=excel_row, column=2).value = people_count
    ws.cell(row=excel_row, column=3).value = company_name
    ws.cell(row=excel_row, column=4).value = question
    ws.cell(row=excel_row, column=5).value = message.chat.id
    ws.cell(row=excel_row, column=6).value = partner_name
    wb.save("TicketsDataBase.xlsx")

    excel_row += 1





while True:
    try:
        bot.polling(none_stop=True)

    except Exception as e:
        print(e)
        time.sleep(15)
