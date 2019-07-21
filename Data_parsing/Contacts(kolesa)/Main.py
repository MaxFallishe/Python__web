from selenium import webdriver
import time
from openpyxl import load_workbook


excel_row = 1
wb = load_workbook('Empty.xlsx')
ws = wb.active

round_num = 1

def main():
    global excel_row
    global round_num

    browser = webdriver.Chrome("D:\\Selenium\\Chrome\\chromedriver.exe")
    browser.maximize_window()

    with open("kolesa(5page_mix).txt") as read_file:
        for url in read_file.readlines():
            browser.get(url)
            try: # допилить условие в браузере последняя ссылка
                button = browser.find_element_by_class_name("offer__contacts").find_element_by_css_selector(".offer__show-phone.action-link.showPhonesLink.js__show-phones")
                button.click()
                phone_number = browser.find_element_by_css_selector(".offer__phones-list.js__phones-list").text
                if phone_number == " ":
                    continue
                    print(round_num)
                try:
                    cities = browser.find_element_by_css_selector(".offer__sidebar.hasNoCredit").find_element_by_class_name("offer__parameters").find_elements_by_tag_name("dl")

                    for i in cities:
                        try:
                            if "Город" in i.text:
                                city = i.text.replace("Город", "")
                                print(city)
                        except:
                            pass

                except Exception as e:
                    print(e)

                ws.cell(row=excel_row, column=1).value = city
                ws.cell(row=excel_row, column=2).value = phone_number

                wb.save('KolesaTable(5pages_mix).xlsx')
                excel_row += 1

            except Exception as e:
                print(e)
                print("Нет номера", url)

            print("round №", round_num)
            round_num += 1
            time.sleep(1)


if __name__ == '__main__':
    main()
