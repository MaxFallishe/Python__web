import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import time


def main():
    excel_row = 1
    wb = load_workbook('EmptyFile.xlsx')
    ws = wb.active

    for i in range(1, 2+1):
        url = generate_url("http://top.uz/trade/orgs/cat3631?page=", str(i), "&sort=0&limit=30")

        page = get_html(url)

        soup = BeautifulSoup(page, "html.parser")
        universities_table = soup.find('td', id='td_content').find_all('table')[-1]
        table_cells = universities_table.find_all('tr')
        # print(universities_table)

        for university_name in table_cells[::2]:
            university_description = table_cells[(table_cells.index(university_name)+1)].text

            # print(university_description)

            description = university_description.split('Адрес:')[0]
            try:
                site = university_description.split("Cайт:")[1]
            except:
                site = 'NO'
            try:
                email = university_description.split("Cайт:")[0].split('E-m@il:')[1]
            except:
                email = 'NO'
            try:
                phones = university_description.split("Cайт:")[0].split('E-m@il:')[0].split('Тел.:')[1]
            except:
                phones = 'NO'
            try:
                address = university_description.split("Cайт:")[0].split('E-m@il:')[0].split('Тел.:')[0].split('Адрес:')[1]
            except:
                address = 'NO'

            excel_arr = [university_name.text, description, site, email, phones, address]

            for j in range(1, 6 + 1):
                ws.cell(row=excel_row, column=j).value = excel_arr[j - 1]
            excel_row += 1



            print("Название", university_name.text)
            print("Описание", description)
            print("Сайт", site)
            print("Почта", email)
            print("Телефоны", phones)
            print("Адресс", address)

            print("---------------------------------------------------")

    wb.save('Colleges.xlsx')


def get_html(url):
    r = requests.get(url)
    return r.text


def generate_url(url_start, num, url_end):
    final_url = url_start + num + url_end
    return final_url


if __name__ == '__main__':
    main()

# .page_source
