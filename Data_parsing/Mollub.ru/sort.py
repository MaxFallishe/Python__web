import openpyxl
import time

book = openpyxl.load_workbook('2.xlsx')
book2 = openpyxl.load_workbook('1.xlsx')

sheet = book.active
sheet2 = book2.active


for i in range(1, 1183+1):

    for j in range(2, 677+1):
        print(sheet.cell(row=i, column=4).value, sheet2.cell(row=j, column=3).value)

        if sheet.cell(row=i, column=4).value == sheet2.cell(row=j, column=3).value:

            sheet.cell(row=i, column=5).value = sheet2.cell(row=j, column=4).value
            sheet.cell(row=i, column=6).value = sheet2.cell(row=j, column=5).value
            sheet.cell(row=i, column=7).value = sheet2.cell(row=j, column=6).value
            sheet.cell(row=i, column=8).value = sheet2.cell(row=j, column=7).value
            sheet.cell(row=i, column=9).value = sheet2.cell(row=j, column=8).value
            sheet.cell(row=i, column=10).value = sheet2.cell(row=j, column=9).value
            sheet.cell(row=i, column=11).value = sheet2.cell(row=j, column=10).value
            sheet.cell(row=i, column=11).value = sheet2.cell(row=j, column=10).value

book.save('F_to_me.xlsx')



