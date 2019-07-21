from selenium import webdriver
import time
from openpyxl import load_workbook


browser = webdriver.Chrome("D:\\Selenium\\Chrome\\chromedriver.exe")
browser.maximize_window()

excel_row = 1
wb = load_workbook('Empty.xlsx')
ws = wb.active

catalog = {
           # "Транспорт": "http://mollub.ru/ru/smazochnye-materialy-i-avtohimija/transport/",
           # "Сельское хозяйство": "http://mollub.ru/ru/smazochnye-materialy-i-avtohimija/sel-skoe-hozajstvo/",
           # "Промышленность": "http://mollub.ru/ru/smazochnye-materialy-i-avtohimija/promy-lennost/",
           "Обсуживание автомобилей": "http://mollub.ru/ru/smazochnye-materialy-i-avtohimija/obsluzivanie-avtomobilej/"
           # "Обсуживание грузовых автомобилей": "http://mollub.ru/ru/smazochnye-materialy-i-avtohimija/obsluzivanie-gruzovyh-avtomobilej/",
           # "Автомобили": "http://mollub.ru/ru/smazochnye-materialy-i-avtohimija/avtomobili/",
           # "Мотоциклы": "http://mollub.ru/ru/smazochnye-materialy-i-avtohimija/motocikly/",
           }

# "Садовое оборудование": "http://mollub.ru/ru/smazochnye-materialy-i-avtohimija/sadovoe-oborudovanie/produkty/"

catalog_name = ""
category_name = ""


def main():
    global catalog_name
    global category_name

    for catalog_key in catalog:

        browser.get(catalog[catalog_key])
        catalog_name = catalog_key
        print(catalog[catalog_key])
        print(catalog_key)
        # print(browser.find_element_by_css_selector(".cf.page_lubricant_row").find_element_by_css_selector('.block1.right_column').text)

        categories = browser.find_element_by_css_selector(".cf.page_lubricant_row")\
            .find_element_by_css_selector('.block1.right_column')\
            .find_element_by_class_name('category_group')\
            .find_elements_by_css_selector('.category_row.anim')

        categories_list = {}

        for category in categories:
            dict_category_link = category.get_attribute('href')
            dict_category_name = category.text

            categories_list[dict_category_name] = dict_category_link

        print(categories_list)
        for category_key in categories_list:
            category_name = category_key
            print(category.text, "<-------------------------------------------------")
            category_link = categories_list[category_key]
            scrap_category_elements(category_link)


def scrap_category_elements(category_link):
    browser.get(category_link)

    elements = browser.find_element_by_css_selector('.lub_product_list_container.cf')\
        .find_element_by_class_name('lub_product_list_items').find_elements_by_tag_name('a')

    elements_list = []
    for element in elements:
        element_link = element.get_attribute('href')
        elements_list.append(element_link)

    for product in elements_list:
        scrap_element_info(product)


def scrap_element_info(product_link):
    global excel_row

    browser.get(product_link)
    try:
        product_title = browser.find_element_by_class_name("lub_product_title").text
    except:
        product_title = "-"

    try:
        product_nomination = browser.find_element_by_class_name("lub_product_nomination_wrapper").text
    except:
        product_nomination = "-"

    try:
        product_description = browser.find_element_by_class_name("lub_product_description_wrapper").text
    except:
        product_description = "-"

    try:
        product_application = browser.find_element_by_class_name("lub_product_applications_wrapper").text
    except:
        product_application = "-"

    try:
        product_features = browser.find_element_by_class_name("lub_product_features_wrapper").text
    except:
        product_features = "-"

    try:
        product_pdf = browser.find_element_by_class_name("download").find_element_by_tag_name('a').get_attribute('href')
    except:
        product_pdf = "-"

    try:
        product_approvals = browser.find_element_by_class_name("lub_table_perf_box").text
    except:
        product_approvals = "-"

    try:
        product_specification = browser.find_element_by_class_name("lub_table_prop_box").text
    except:
        product_specification = "-"

    print('----------------------------------------------')
    print(catalog_name)
    print(category_name)
    print(product_title)
    print(product_nomination)
    print(product_description)
    print(product_application)
    print(product_features)
    print(product_pdf)
    print(product_approvals)
    print(product_specification)

    ws.cell(row=excel_row, column=1).value = catalog_name
    ws.cell(row=excel_row, column=2).value = category_name
    ws.cell(row=excel_row, column=3).value = product_title
    ws.cell(row=excel_row, column=4).value = product_nomination
    ws.cell(row=excel_row, column=5).value = product_description
    ws.cell(row=excel_row, column=6).value = product_application
    ws.cell(row=excel_row, column=7).value = product_features
    ws.cell(row=excel_row, column=8).value = product_pdf
    ws.cell(row=excel_row, column=9).value = product_approvals
    ws.cell(row=excel_row, column=10).value = product_specification

    wb.save('autoServiceOnly.xlsx')
    excel_row += 1


if __name__ == '__main__':
    main()












