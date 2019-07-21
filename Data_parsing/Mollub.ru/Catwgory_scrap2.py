from selenium import webdriver
from openpyxl import load_workbook
import time


browser = webdriver.Chrome("D:\\Selenium\\Chrome\\chromedriver.exe")
browser.maximize_window()


excel_row = 1
wb = load_workbook('Empty.xlsx')
ws = wb.active


def main():
    global excel_row

    browser.get("http://mollub.ru/ru/smazochnye-materialy-i-avtohimija/obsluzivanie-avtomobilej/")

    categories = browser.find_element_by_css_selector(".cf.page_lubricant_row") \
        .find_element_by_css_selector('.block1.right_column') \
        .find_element_by_class_name('category_group') \
        .find_elements_by_css_selector('.category_row.anim')

    categories_links = []
    categories_names = []

    for category in categories:
        categories_links.append(category.get_attribute('href'))
        categories_names.append(category.text)

    print(categories_names)
    print(categories_links)

    for i in range(len(categories_links)):
        browser.get(categories_links[i])
        print('_______', categories_names[i], '_______')

        groups = browser.find_element_by_class_name('lub_sub_category_list').find_element_by_class_name('active').find_elements_by_tag_name('li')

        if groups ==[]:
            print('Empty')
            elements = browser.find_element_by_class_name('lub_product_list_items').find_elements_by_css_selector('.filtered')
            for element in elements:

                print(element.find_element_by_tag_name('h2').text)
                ws.cell(row=excel_row, column=1).value = categories_names[i]
                ws.cell(row=excel_row, column=3).value = element.find_element_by_tag_name('h2').text

                wb.save('Test2.xlsx')
                excel_row += 1
            continue

        for group in groups:
            group.click()
            time.sleep(2)
            print('---------', group.text, '---------')

            elements = browser.find_element_by_class_name('lub_product_list_items').find_elements_by_css_selector('.filtered')
            for element in elements:
                print(element.find_element_by_tag_name('h2').text)

                ws.cell(row=excel_row, column=1).value = categories_names[i]
                ws.cell(row=excel_row, column=2).value = group.text
                ws.cell(row=excel_row, column=3).value = element.find_element_by_tag_name('h2').text

                wb.save('Test2.xlsx')
                excel_row += 1

            group.click()

        time.sleep(2)


if __name__ == '__main__':
    main()
