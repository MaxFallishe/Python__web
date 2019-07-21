from selenium import webdriver
from openpyxl import load_workbook


browser = webdriver.Chrome("D:\\Selenium\\Chrome\\chromedriver.exe")
browser.maximize_window()


excel_row = 173
wb = load_workbook('Kost.xlsx')
ws = wb.active


def main():
    global excel_row

    browser.get('http://mollub.ru/ru/smazochnye-materialy-i-avtohimija/avtomobili/produkty/ohlazdausie-zidkosti-s-ponizennoj-temperaturoj-zamerzania.bolshie-servisnye-intervaly /')

    items = browser.find_elements_by_css_selector('.bolshie-servisnye-intervaly ')

    print(len(items))

    for i in items:
        name = i.find_element_by_tag_name("h2").text

        if name == "":
            print("empty")
            continue

        print(name)
        ws.cell(row=excel_row, column=1).value = "Автомобили"
        ws.cell(row=excel_row, column=2).value = "Охлаждающие жидкости с пониженной температурой замерзания"
        ws.cell(row=excel_row, column=3).value = "Большие сервисные интервалы"
        ws.cell(row=excel_row, column=4).value = name

        excel_row += 1

    wb.save('Kost.xlsx')


if __name__ == '__main__':
    main()
