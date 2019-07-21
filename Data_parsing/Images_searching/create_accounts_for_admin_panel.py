from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from openpyxl import load_workbook
import time


# excel_row = 1
wb = load_workbook('UzInstitutes.xlsx')
ws = wb.active

browser = webdriver.Chrome("D:\\Selenium\\Chrome\\chromedriver.exe")
browser.maximize_window()


def main():
    authorization()
    for i in range(58, 82+1):
        name = ws.cell(row=i, column=2).value
        email = ws.cell(row=i, column=11).value
        add_institute(name, email)
        print("added element number", i)
        time.sleep(10)


def authorization():
    load_page_and_wait("https://talaba-9d251.firebaseapp.com/auth", 1)

    login = browser.find_elements_by_css_selector('.ivu-input.ivu-input-default')[0]
    password = browser.find_elements_by_css_selector('.ivu-input.ivu-input-default')[1]
    enter_button = browser.find_element_by_css_selector('.ivu-btn.ivu-btn-primary.ivu-btn-long')

    login.send_keys('admin@gmail.com')
    password.send_keys('test123')
    enter_button.click()


def add_institute(name, email):
    time.sleep(4)
    add_button = browser.find_element_by_class_name('ivu-card-extra')
    add_button.click()

    institute_name = browser.find_elements_by_css_selector('.ivu-input.ivu-input-default')[0]
    institute_email = browser.find_elements_by_css_selector('.ivu-input.ivu-input-default')[1]
    ok_button = browser.find_element_by_css_selector('.ivu-btn.ivu-btn-primary.ivu-btn-large')

    institute_name.send_keys(name)
    institute_email.send_keys(email)
    ok_button.click()


def load_page_and_wait(link, waiting_time):
    browser.get(link)
    time.sleep(waiting_time)


if __name__ == '__main__':
    main()




