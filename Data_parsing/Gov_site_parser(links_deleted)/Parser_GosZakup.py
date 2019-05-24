from selenium import webdriver
import time
from openpyxl import load_workbook

url = "xxxxxxxxxxxxxxxxxxxxxxxx"
count = 2  # need for correct writing in excel


def main():
    browser = webdriver.Chrome('chromedriver.exe')
    browser.maximize_window()
    browser2 = webdriver.Chrome('chromedriver.exe')
    browser.maximize_window()

    browser.get(url)
    total_page = get_total_page(browser)
    print(total_page)

    wb = load_workbook('Main_sample.xlsx')
    ws = wb.active

    for i in range(1, total_page + 1):
        print(url + str(i))
        browser.get(url + str(i))
        scraper(browser, browser2, ws)
    
    wb.save('GERBS_finalTable(April).xlsx')


def get_total_page(browser):
    page_selector = browser.find_element_by_css_selector('.col-md-6.text-center').find_elements_by_tag_name('li')
    total_page = page_selector[-1].find_element_by_tag_name('a').get_attribute('href').split('&page=')[-1]
    return int(total_page)


def scraper(browser, browser2, ws):
    global count
    inf = browser.find_element_by_id('search-result_wrapper').find_element_by_id('search-result').\
        find_element_by_tag_name('tbody').find_elements_by_tag_name('tr')

    for i in inf:
        attributes = i.find_elements_by_tag_name('td')
        order = attributes[0].text
        customer = attributes[1].text
        name = attributes[2].text
        purchase_method = attributes[3].text
        unit = attributes[4].text
        quantity = attributes[5].text
        price_per_one = attributes[6].text
        planned_amount = attributes[7].text
        purchase_term = attributes[8].text
        status = attributes[9].text

        link_customer = attributes[1].find_element_by_tag_name('a').get_attribute('href')
        link_name = attributes[2].find_element_by_tag_name('a').get_attribute('href')

        browser2.get(link_customer)
        customer_table = browser2.find_element_by_css_selector('.table.table-striped').find_elements_by_tag_name('tr')

        email = 0
        full_name = 0
        phone = 0
        specification = 0

        for j in customer_table:
            if j.find_element_by_tag_name('th').text == 'E-Mail:':
                email = j.find_element_by_tag_name('td').text
            elif j.find_element_by_tag_name('th').text == 'Контактный телефон:':
                phone = j.find_element_by_tag_name('td').text

        last_table = browser2.find_elements_by_css_selector('.table.table-striped')
        manager = last_table[-2].find_elements_by_tag_name('tr')

        for j in manager:
            if j.find_element_by_tag_name('th').text == 'ФИО':
                full_name = j.find_element_by_tag_name('td').text

        browser2.get(link_name)
        feature_1 = ''
        feature_2 = ''
        for j in browser2.find_elements_by_css_selector('.panel.panel-default'):
            j_list = j.text.split('\n')
            if j_list[0] == 'Описание товары, работы, услуги':
                feature_1 = j_list[11]
                feature_2 = j_list[15]
                specification = feature_1 + ' ' + feature_2
                break

        # print(order)
        # print(customer)
        # print(name)
        # print(purchase_method)
        # print(unit)
        # print(quantity)
        # print(price_per_one)
        # print(planned_amount)
        # print(purchase_term)
        # print(status)
        # print(link_customer)
        # print(link_name)

        excel_arr = [order, customer, email, full_name, phone, name, specification, purchase_method, unit, quantity,
                     price_per_one, planned_amount, purchase_term, status]
        print(excel_arr)
        for j in range(0, len(excel_arr)):
            if excel_arr[j] == '':
                excel_arr[j] = 0
        for j in range(1, 14 + 1):
            ws.cell(row=count, column=j).value = excel_arr[j - 1]

        count += 1
        # time.sleep(2000)

    #order = 1
    #customer = 1
    #email = 1
    #full_name = 1
    #phone = 1
    #name = 1
    #specification = 1
    #purchase_method = 1
    #unit = 1
    #quantity = 1
    #price_per_one = 1
    #planned_amount = 1
    #purchase_term = 1
    #status = 1


if __name__ == '__main__':
    main()



