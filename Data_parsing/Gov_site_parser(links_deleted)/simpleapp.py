# -*- coding: utf-8 -*-
from PyQt5 import QtCore, QtGui, QtWidgets
from selenium import webdriver
import time
from openpyxl import load_workbook
count = 2  # need for correct writing in excel


def main(url, file_name):
    browser = webdriver.Chrome('chromedriver.exe')
    browser.maximize_window()
    browser2 = webdriver.Chrome('chromedriver.exe')
    browser.maximize_window()

    browser.get("https://" + url)
    total_page = get_total_page(browser)
    print(total_page)
    wb = load_workbook('Main_sample.xlsx')
    ws = wb.active

    for i in range(1, total_page + 1):

        print(url + "&page=" + str(i))
        browser.get("https://" + url + "&page=" + str(i))

        scraper(browser, browser2, ws)

    if file_name != "":
        wb.save(str(file_name) + ".xlsx")
    else:
        wb.save("Excel_Table.xlsx")


def get_total_page(browser):
    page_selector = browser.find_element_by_css_selector('.col-md-6.text-center').find_elements_by_tag_name('li')
    total_page = page_selector[-1].find_element_by_tag_name('a').get_attribute('href').split('&page=')[-1]
    return int(total_page)


def scraper(browser, browser2, ws):
    global count
    inf = browser.find_element_by_id('search-result_wrapper').find_element_by_id('search-result').\
        find_element_by_tag_name('tbody').find_elements_by_tag_name('tr')

    for i in inf:
        attributes = i.find_elements_by_tag_name('td')
        order = attributes[0].text
        customer = attributes[1].text
        name = attributes[2].text
        purchase_method = attributes[3].text
        unit = attributes[4].text
        quantity = attributes[5].text
        price_per_one = attributes[6].text
        planned_amount = attributes[7].text
        purchase_term = attributes[8].text
        status = attributes[9].text

        link_customer = attributes[1].find_element_by_tag_name('a').get_attribute('href')
        link_name = attributes[2].find_element_by_tag_name('a').get_attribute('href')

        browser2.get(link_customer)
        customer_table = browser2.find_element_by_css_selector('.table.table-striped').find_elements_by_tag_name('tr')

        email = 0
        full_name = 0
        phone = 0
        specification = 0

        for j in customer_table:
            if j.find_element_by_tag_name('th').text == 'E-Mail:':
                email = j.find_element_by_tag_name('td').text
            elif j.find_element_by_tag_name('th').text == 'Контактный телефон:':
                phone = j.find_element_by_tag_name('td').text

        last_table = browser2.find_elements_by_css_selector('.table.table-striped')
        manager = last_table[-2].find_elements_by_tag_name('tr')

        for j in manager:
            if j.find_element_by_tag_name('th').text == 'ФИО':
                full_name = j.find_element_by_tag_name('td').text

        browser2.get(link_name)
        feature_1 = ''
        feature_2 = ''
        for j in browser2.find_elements_by_css_selector('.panel.panel-default'):
            j_list = j.text.split('\n')
            if j_list[0] == 'Описание товары, работы, услуги':
                feature_1 = j_list[11]
                feature_2 = j_list[15]
                specification = feature_1 + ' ' + feature_2
                break

        excel_arr = [order, customer, email, full_name, phone, name, specification, purchase_method, unit, quantity,
                     price_per_one, planned_amount, purchase_term, status]
        print(excel_arr)
        for j in range(0, len(excel_arr)):
            if excel_arr[j] == '':
                excel_arr[j] = 0
        for j in range(1, 14 + 1):
            ws.cell(row=count, column=j).value = excel_arr[j - 1]

        count += 1


class Ui_Dialog(object):
    def parsingStart(self):
        print("Button was click")
        name = self.textEdit.toPlainText()
        customer = self.textEdit_2.toPlainText()
        number_pp = self.textEdit_3.toPlainText()
        specification = self.comboBox_2.currentText()
        status = self.comboBox_3.currentText()
        purchase_date = self.comboBox_5.currentText()
        purchase_method = self.comboBox_4.currentText()
        purchase_subject = self.comboBox_6.currentText()
        subject_sigh = self.comboBox_7.currentText()
        year = self.textEdit_4.toPlainText()

        specification_dict = {
            "Ничего не выбрано": "",
            "Бюджетные кредиты специализированным организациям": "925",
            "Взносы на обязательное страхование": "878",
            "Возврат трансфертов общего характера в случаях, предусмотренных бюджетным законодательством": "957",
            "Затраты Фонда всеобщего обязательного среднего образования": "896",
            "Капитальный ремонт дорог": "916",
            "Капитальный ремонт помещений, зданий, сооружений государственных предприятий": "917",
            "Капитальный ремонт помещений, зданий, сооружений, передаточных устройств": "862",
            "Капитальный ремонт прочих основных средств": "918",
            "Командировки и служебные разъезды внутри страны": "853",
            "Командировки и служебные разъезды внутри страны технического персонала": "883",
            "Командировки и служебные разъезды за пределы страны": "895",
            "Материально-техническое оснащение государственных предприятий": "914",
            "Обязательные профессиональные пенсионные взносы": "3405",
            "Оплата аренды за помещение": "891",
            "Оплата коммунальных услуг": "852",
            "Оплата консалтинговых услуг и исследований": "893",
            "Оплата прочих услуг и работ": "894",
            "Оплата транспортных услуг": "890",
            "Оплата труда": "848",
            "Оплата труда технического персонала": "850",
            "Оплата услуг в рамках государственного социального заказа": "892",
            "Оплата услуг связи": "889"
        }

        status_dict = {
            "Ничего не выбрано": "",
            "Договор действует": "16",
            "Договор изменен": "320",
            "Договор не заключен": "18",
            "Договор не заключен, отказ от первого победителя": "400",
            "Договор не заключен. Изменен.": "350",
            "Договор не исполнен. Нет платежей": "340",
            "Договор частично исполнен": "410",
            "Закупка не состоялась": "10",
            "Закупка не состоялась. Изменен": "11",
            "Закупка состоялась": "9",
            "Заявка": "3",
            "Заявка возвращена организатором": "13",
            "Заявка отклонена": "27",
            "Заявка подтверждена": "26",
            "Изменен": "6",
            "Исполнен": "19",
            "На обжаловании": "22",
            "Опубликован": "5",
            "Отказ от закупки": "20",
            "Отменен": "7",
            "Отменен. Изменен": "21",
            "Передан": "24",
            "Пересмотр итогов": "23",
            "Принятие решение об исполнении уведомления": "540",
            "Приостановлен": "12",
            "Проект договора": "310",
            "Проект заявки": "25",
            "Проект лота": "4",
            "Утвержден": "2"
        }

        purchase_method_dict = {
            "Ничего не выбрано": "",
            "Запрос ценовых предложений": "3",
            "Открытый конкурс": "2",
            "Аукцион": "7",
            "Из одного источника по несостоявшимся закупкам": "6",
            "Через товарные биржи": "8",
            "Из одного источника путем прямого заключения договора": "23",
            "Закупка жилища": "50",
            "Закупка по государственному социальному заказу": "52",
            "Конкурс с предварительным квалификационным отбором": "32",
            "Конкурс с применением двухэтапных процедур": "22",
            "Из одного источника (не ГЗ)": "105",
            "На организованных электронных торгах (не ГЗ)": "107",
            "Запрос ценовых предложений (не ГЗ)": "116",
            "Тендер (не ГЗ)": "117",
            "Через товарные биржи (не ГЗ)": "118"
        }

        purchase_date_dict = {
            "Ничего не выбрано": "",
            "Январь": "1",
            "Февраль": "2",
            "Март": "3",
            "Апрель": "4",
            "Май": "5",
            "Июнь": "6",
            "Июль": "7",
            "Август": "8",
            "Сентябрь": "9",
            "Октябрь": "10",
            "Ноябрь": "11",
            "Декабрь": "12"

        }

        purchase_subject_dict = {
            "Ничего не выбрано": "",
            "Товар": "G",
            "Услуга": "S",
            "Работа": "R"
        }

        subject_sigh_dict = {
            "Ничего не выбрано": "",
            "Субъект ГЗ": "1",
            "Не суюъект ГЗ": "2"
        }

        # Many conditions below
        start_url = "xxxxxxxxxxxxxxxxxxxxxxxxxxxxxx"
        work_url = start_url + name
        print("text is:" + purchase_subject_dict[purchase_subject] + ":")
        if customer != "":
            work_url = work_url + "&search=&filter%5Bcustomer%5D=" + customer
        if purchase_method_dict[purchase_method] != "":
            work_url = work_url + "&filter%5Bmethod%5D%5B%5D=" + purchase_method_dict[purchase_method]
        if purchase_date_dict[purchase_date] != "":
            work_url = work_url + "&filter%5Bmonth%5D%5B%5D=" + purchase_date_dict[purchase_date]
        if year != "":
            work_url = work_url + "&filter%5Byear%5D%5B%5D=" + year
        if specification_dict[specification] != "":
            work_url = work_url + "&filter%5Bspec%5D=" + specification_dict[specification]
        if status_dict[status] != "":
            work_url = work_url + "&filter%5Bstatus%5D=" + status_dict[status]
        if number_pp != "":
            work_url = work_url + "&filter%5Bnumber%5D=" + number_pp
        if purchase_subject_dict[purchase_subject] != "":
            work_url = work_url + "&filter%5Bsubject_type%5D%5B%5D=" + purchase_subject_dict[purchase_subject]
        if subject_sigh_dict[subject_sigh] != "":
            work_url = work_url + "&filter%5Bqvazi%5D=" + subject_sigh_dict[subject_sigh]

        print(work_url)
        main(work_url,self.textEdit_5.toPlainText())

    def setupUi(self, Dialog):
        Dialog.setObjectName("Dialog")
        Dialog.resize(590, 566)
        self.label = QtWidgets.QLabel(Dialog)
        self.label.setGeometry(QtCore.QRect(150, -20, 621, 121))
        font = QtGui.QFont()
        font.setPointSize(20)
        font.setBold(False)
        font.setUnderline(True)
        font.setWeight(50)
        self.label.setFont(font)
        self.label.setObjectName("label")
        self.label_2 = QtWidgets.QLabel(Dialog)
        self.label_2.setGeometry(QtCore.QRect(10, 80, 201, 20))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label_2.setFont(font)
        self.label_2.setObjectName("label_2")
        self.label_3 = QtWidgets.QLabel(Dialog)
        self.label_3.setGeometry(QtCore.QRect(10, 150, 201, 20))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label_3.setFont(font)
        self.label_3.setObjectName("label_3")
        self.comboBox_2 = QtWidgets.QComboBox(Dialog)
        self.comboBox_2.setGeometry(QtCore.QRect(10, 170, 271, 22))
        self.comboBox_2.setObjectName("comboBox_2")
        self.comboBox_2.addItem("")
        self.comboBox_2.addItem("")
        self.comboBox_2.addItem("")
        self.comboBox_2.addItem("")
        self.comboBox_2.addItem("")
        self.comboBox_2.addItem("")
        self.comboBox_2.addItem("")
        self.comboBox_2.addItem("")
        self.comboBox_2.addItem("")
        self.comboBox_2.addItem("")
        self.comboBox_2.addItem("")
        self.comboBox_2.addItem("")
        self.comboBox_2.addItem("")
        self.comboBox_2.addItem("")
        self.comboBox_2.addItem("")
        self.comboBox_2.addItem("")
        self.comboBox_2.addItem("")
        self.comboBox_2.addItem("")
        self.comboBox_2.addItem("")
        self.comboBox_2.addItem("")
        self.comboBox_2.addItem("")
        self.comboBox_2.addItem("")
        self.comboBox_2.addItem("")
        self.textEdit = QtWidgets.QTextEdit(Dialog)
        self.textEdit.setGeometry(QtCore.QRect(10, 100, 271, 31))
        self.textEdit.setObjectName("textEdit")
        self.label_4 = QtWidgets.QLabel(Dialog)
        self.label_4.setGeometry(QtCore.QRect(300, 80, 201, 20))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label_4.setFont(font)
        self.label_4.setObjectName("label_4")
        self.textEdit_2 = QtWidgets.QTextEdit(Dialog)
        self.textEdit_2.setGeometry(QtCore.QRect(300, 100, 271, 31))
        self.textEdit_2.setObjectName("textEdit_2")
        self.textEdit_5 = QtWidgets.QTextEdit(Dialog)
        self.textEdit_5.setGeometry(QtCore.QRect(300, 380, 271, 31))
        self.textEdit_5.setObjectName("textEdit_5")
        self.label_12 = QtWidgets.QLabel(Dialog)
        self.label_12.setGeometry(QtCore.QRect(300, 360, 201, 20))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label_12.setFont(font)
        self.label_12.setObjectName("label_12")
        self.comboBox_3 = QtWidgets.QComboBox(Dialog)
        self.comboBox_3.setGeometry(QtCore.QRect(10, 240, 271, 22))
        self.comboBox_3.setObjectName("comboBox_3")
        self.comboBox_3.addItem("")
        self.comboBox_3.addItem("")
        self.comboBox_3.addItem("")
        self.comboBox_3.addItem("")
        self.comboBox_3.addItem("")
        self.comboBox_3.addItem("")
        self.comboBox_3.addItem("")
        self.comboBox_3.addItem("")
        self.comboBox_3.addItem("")
        self.comboBox_3.addItem("")
        self.comboBox_3.addItem("")
        self.comboBox_3.addItem("")
        self.comboBox_3.addItem("")
        self.comboBox_3.addItem("")
        self.comboBox_3.addItem("")
        self.comboBox_3.addItem("")
        self.comboBox_3.addItem("")
        self.comboBox_3.addItem("")
        self.comboBox_3.addItem("")
        self.comboBox_3.addItem("")
        self.comboBox_3.addItem("")
        self.comboBox_3.addItem("")
        self.comboBox_3.addItem("")
        self.comboBox_3.addItem("")
        self.comboBox_3.addItem("")
        self.comboBox_3.addItem("")
        self.comboBox_3.addItem("")
        self.comboBox_3.addItem("")
        self.comboBox_3.addItem("")
        self.comboBox_3.addItem("")
        self.label_5 = QtWidgets.QLabel(Dialog)
        self.label_5.setGeometry(QtCore.QRect(10, 220, 201, 20))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label_5.setFont(font)
        self.label_5.setObjectName("label_5")
        self.label_6 = QtWidgets.QLabel(Dialog)
        self.label_6.setGeometry(QtCore.QRect(300, 210, 201, 20))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label_6.setFont(font)
        self.label_6.setObjectName("label_6")
        self.textEdit_3 = QtWidgets.QTextEdit(Dialog)
        self.textEdit_3.setGeometry(QtCore.QRect(300, 230, 271, 31))
        self.textEdit_3.setObjectName("textEdit_3")
        self.label_7 = QtWidgets.QLabel(Dialog)
        self.label_7.setGeometry(QtCore.QRect(10, 290, 201, 20))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label_7.setFont(font)
        self.label_7.setObjectName("label_7")
        self.comboBox_4 = QtWidgets.QComboBox(Dialog)
        self.comboBox_4.setGeometry(QtCore.QRect(10, 310, 271, 22))
        self.comboBox_4.setObjectName("comboBox_4")
        self.comboBox_4.addItem("")
        self.comboBox_4.addItem("")
        self.comboBox_4.addItem("")
        self.comboBox_4.addItem("")
        self.comboBox_4.addItem("")
        self.comboBox_4.addItem("")
        self.comboBox_4.addItem("")
        self.comboBox_4.addItem("")
        self.comboBox_4.addItem("")
        self.comboBox_4.addItem("")
        self.comboBox_4.addItem("")
        self.comboBox_4.addItem("")
        self.comboBox_4.addItem("")
        self.comboBox_4.addItem("")
        self.comboBox_4.addItem("")
        self.comboBox_5 = QtWidgets.QComboBox(Dialog)
        self.comboBox_5.setGeometry(QtCore.QRect(10, 370, 271, 22))
        self.comboBox_5.setObjectName("comboBox_5")
        self.comboBox_5.addItem("")
        self.comboBox_5.addItem("")
        self.comboBox_5.addItem("")
        self.comboBox_5.addItem("")
        self.comboBox_5.addItem("")
        self.comboBox_5.addItem("")
        self.comboBox_5.addItem("")
        self.comboBox_5.addItem("")
        self.comboBox_5.addItem("")
        self.comboBox_5.addItem("")
        self.comboBox_5.addItem("")
        self.comboBox_5.addItem("")
        self.label_8 = QtWidgets.QLabel(Dialog)
        self.label_8.setGeometry(QtCore.QRect(10, 350, 201, 20))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label_8.setFont(font)
        self.label_8.setObjectName("label_8")
        self.comboBox_6 = QtWidgets.QComboBox(Dialog)
        self.comboBox_6.setGeometry(QtCore.QRect(10, 430, 271, 22))
        self.comboBox_6.setObjectName("comboBox_6")
        self.comboBox_6.addItem("")
        self.comboBox_6.addItem("")
        self.comboBox_6.addItem("")
        self.comboBox_6.addItem("")
        self.label_9 = QtWidgets.QLabel(Dialog)
        self.label_9.setGeometry(QtCore.QRect(10, 410, 201, 20))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label_9.setFont(font)
        self.label_9.setObjectName("label_9")
        self.textEdit_4 = QtWidgets.QTextEdit(Dialog)
        self.textEdit_4.setGeometry(QtCore.QRect(300, 160, 271, 31))
        self.textEdit_4.setObjectName("textEdit_4")
        self.label_10 = QtWidgets.QLabel(Dialog)
        self.label_10.setGeometry(QtCore.QRect(300, 140, 201, 20))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label_10.setFont(font)
        self.label_10.setObjectName("label_10")
        self.comboBox_7 = QtWidgets.QComboBox(Dialog)
        self.comboBox_7.setGeometry(QtCore.QRect(10, 490, 271, 22))
        self.comboBox_7.setObjectName("comboBox_7")
        self.comboBox_7.addItem("")
        self.comboBox_7.addItem("")
        self.comboBox_7.addItem("")
        self.label_11 = QtWidgets.QLabel(Dialog)
        self.label_11.setGeometry(QtCore.QRect(10, 470, 201, 20))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label_11.setFont(font)
        self.label_11.setObjectName("label_11")
        self.pushButton = QtWidgets.QPushButton(Dialog)
        self.pushButton.setGeometry(QtCore.QRect(300, 420, 271, 91))
        self.pushButton.setObjectName("pushButton")
        self.pushButton.clicked.connect(self.parsingStart)
        self.retranslateUi(Dialog)
        QtCore.QMetaObject.connectSlotsByName(Dialog)

    def retranslateUi(self, Dialog):
        _translate = QtCore.QCoreApplication.translate
        Dialog.setWindowTitle(_translate("Dialog", "Dialog"))
        self.label.setText(_translate("Dialog", "Парсинг сайта goszakup"))
        self.label_2.setText(_translate("Dialog", "Наименование пункта плана"))
        self.label_3.setText(_translate("Dialog", "Спецификация"))
        self.comboBox_2.setItemText(0, _translate("Dialog", "Ничего не выбрано"))
        self.comboBox_2.setItemText(1, _translate("Dialog", "Бюджетные кредиты специализированным организациям"))
        self.comboBox_2.setItemText(2, _translate("Dialog", "Взносы на обязательное страхование"))
        self.comboBox_2.setItemText(3, _translate("Dialog", "Возврат трансфертов общего характера в случаях, предусмотренных бюджетным законодательством"))
        self.comboBox_2.setItemText(4, _translate("Dialog", "Затраты Фонда всеобщего обязательного среднего образования"))
        self.comboBox_2.setItemText(5, _translate("Dialog", "Капитальный ремонт дорог"))
        self.comboBox_2.setItemText(6, _translate("Dialog", "Капитальный ремонт помещений, зданий, сооружений государственных предприятий"))
        self.comboBox_2.setItemText(7, _translate("Dialog", "Капитальный ремонт помещений, зданий, сооружений, передаточных устройств"))
        self.comboBox_2.setItemText(8, _translate("Dialog", "Капитальный ремонт прочих основных средств"))
        self.comboBox_2.setItemText(9, _translate("Dialog", "Командировки и служебные разъезды внутри страны"))
        self.comboBox_2.setItemText(10, _translate("Dialog", "Командировки и служебные разъезды внутри страны технического персонала"))
        self.comboBox_2.setItemText(11, _translate("Dialog", "Командировки и служебные разъезды за пределы страны"))
        self.comboBox_2.setItemText(12, _translate("Dialog", "Материально-техническое оснащение государственных предприятий"))
        self.comboBox_2.setItemText(13, _translate("Dialog", "Обязательные профессиональные пенсионные взносы"))
        self.comboBox_2.setItemText(14, _translate("Dialog", "Оплата аренды за помещение"))
        self.comboBox_2.setItemText(15, _translate("Dialog", "Оплата коммунальных услуг"))
        self.comboBox_2.setItemText(16, _translate("Dialog", "Оплата консалтинговых услуг и исследований"))
        self.comboBox_2.setItemText(17, _translate("Dialog", "Оплата прочих услуг и работ"))
        self.comboBox_2.setItemText(18, _translate("Dialog", "Оплата транспортных услуг"))
        self.comboBox_2.setItemText(19, _translate("Dialog", "Оплата труда"))
        self.comboBox_2.setItemText(20, _translate("Dialog", "Оплата труда технического персонала"))
        self.comboBox_2.setItemText(21, _translate("Dialog", "Оплата услуг в рамках государственного социального заказа"))
        self.comboBox_2.setItemText(22, _translate("Dialog", "Оплата услуг связи"))
        self.label_4.setText(_translate("Dialog", "Наименование заказчика"))
        self.label_12.setText(_translate("Dialog", "Имя для файла excel "))
        self.comboBox_3.setItemText(0, _translate("Dialog", "Ничего не выбрано"))
        self.comboBox_3.setItemText(1, _translate("Dialog", "Договор действует"))
        self.comboBox_3.setItemText(2, _translate("Dialog", "Договор изменен"))
        self.comboBox_3.setItemText(3, _translate("Dialog", "Договор не заключен"))
        self.comboBox_3.setItemText(4, _translate("Dialog", "Договор не заключен, отказ от первого победителя"))
        self.comboBox_3.setItemText(5, _translate("Dialog", "Договор не заключен. Изменен."))
        self.comboBox_3.setItemText(6, _translate("Dialog", "Договор не исполнен. Нет платежей"))
        self.comboBox_3.setItemText(7, _translate("Dialog", "Договор частично исполнен"))
        self.comboBox_3.setItemText(8, _translate("Dialog", "Закупка не состоялась"))
        self.comboBox_3.setItemText(9, _translate("Dialog", "Закупка не состоялась. Изменен"))
        self.comboBox_3.setItemText(10, _translate("Dialog", "Закупка состоялась"))
        self.comboBox_3.setItemText(11, _translate("Dialog", "Заявка"))
        self.comboBox_3.setItemText(12, _translate("Dialog", "Заявка возвращена организатором"))
        self.comboBox_3.setItemText(13, _translate("Dialog", "Заявка отклонена"))
        self.comboBox_3.setItemText(14, _translate("Dialog", "Заявка подтверждена"))
        self.comboBox_3.setItemText(15, _translate("Dialog", "Изменен"))
        self.comboBox_3.setItemText(16, _translate("Dialog", "Исполнен"))
        self.comboBox_3.setItemText(17, _translate("Dialog", "На обжаловании"))
        self.comboBox_3.setItemText(18, _translate("Dialog", "Опубликован"))
        self.comboBox_3.setItemText(19, _translate("Dialog", "Отказ от закупки"))
        self.comboBox_3.setItemText(20, _translate("Dialog", "Отменен"))
        self.comboBox_3.setItemText(21, _translate("Dialog", "Отменен. Изменен"))
        self.comboBox_3.setItemText(22, _translate("Dialog", "Передан"))
        self.comboBox_3.setItemText(23, _translate("Dialog", "Пересмотр итогов"))
        self.comboBox_3.setItemText(24, _translate("Dialog", "Принятие решение об исполнении уведомления"))
        self.comboBox_3.setItemText(25, _translate("Dialog", "Приостановлен"))
        self.comboBox_3.setItemText(26, _translate("Dialog", "Проект договора"))
        self.comboBox_3.setItemText(27, _translate("Dialog", "Проект заявки"))
        self.comboBox_3.setItemText(28, _translate("Dialog", "Проект лота"))
        self.comboBox_3.setItemText(29, _translate("Dialog", "Утвержден"))
        self.label_5.setText(_translate("Dialog", "Статус"))
        self.label_6.setText(_translate("Dialog", "Номер п/п"))
        self.label_7.setText(_translate("Dialog", "Способ закупки"))
        self.comboBox_4.setItemText(0, _translate("Dialog", "Ничего не выбрано"))
        self.comboBox_4.setItemText(1, _translate("Dialog", "Запрос ценовых предложений"))
        self.comboBox_4.setItemText(2, _translate("Dialog", "Открытый конкурс"))
        self.comboBox_4.setItemText(3, _translate("Dialog", "Аукцион"))
        self.comboBox_4.setItemText(4, _translate("Dialog", "Из одного источника по несостоявшимся закупкам"))
        self.comboBox_4.setItemText(5, _translate("Dialog", "Через товарные биржи"))
        self.comboBox_4.setItemText(6, _translate("Dialog", "Из одного источника путем прямого заключения договора"))
        self.comboBox_4.setItemText(7, _translate("Dialog", "Закупка жилища"))
        self.comboBox_4.setItemText(8, _translate("Dialog", "Закупка по государственному социальному заказу"))
        self.comboBox_4.setItemText(9, _translate("Dialog", "Конкурс с предварительным квалификационным отбором"))
        self.comboBox_4.setItemText(10, _translate("Dialog", "Конкурс с применением двухэтапных процедур"))
        self.comboBox_4.setItemText(11, _translate("Dialog", "Из одного источника (не ГЗ)"))
        self.comboBox_4.setItemText(12, _translate("Dialog", "На организованных электронных торгах (не ГЗ)"))
        self.comboBox_4.setItemText(13, _translate("Dialog", "Запрос ценовых предложений (не ГЗ)"))
        self.comboBox_4.setItemText(14, _translate("Dialog", "Тендер (не ГЗ)"))
        self.comboBox_4.setItemText(15, _translate("Dialog", "Через товарные биржи (не ГЗ)"))
        self.comboBox_5.setItemText(0, _translate("Dialog", "Ничего не выбрано"))
        self.comboBox_5.setItemText(1, _translate("Dialog", "Январь"))
        self.comboBox_5.setItemText(2, _translate("Dialog", "Февраль"))
        self.comboBox_5.setItemText(3, _translate("Dialog", "Март"))
        self.comboBox_5.setItemText(4, _translate("Dialog", "Апрель"))
        self.comboBox_5.setItemText(5, _translate("Dialog", "Май"))
        self.comboBox_5.setItemText(6, _translate("Dialog", "Июнь"))
        self.comboBox_5.setItemText(7, _translate("Dialog", "Июль"))
        self.comboBox_5.setItemText(8, _translate("Dialog", "Август"))
        self.comboBox_5.setItemText(9, _translate("Dialog", "Сентябрь"))
        self.comboBox_5.setItemText(10, _translate("Dialog", "Октябрь"))
        self.comboBox_5.setItemText(11, _translate("Dialog", "Ноябрь"))
        self.comboBox_5.setItemText(12, _translate("Dialog", "Декабрь"))
        self.label_8.setText(_translate("Dialog", "Срок закупки"))
        self.comboBox_6.setItemText(0, _translate("Dialog", "Ничего не выбрано"))
        self.comboBox_6.setItemText(1, _translate("Dialog", "Товар"))
        self.comboBox_6.setItemText(2, _translate("Dialog", "Услуга"))
        self.comboBox_6.setItemText(3, _translate("Dialog", "Работа"))
        self.label_9.setText(_translate("Dialog", "Предмет закупки"))
        self.label_10.setText(_translate("Dialog", "Финансовый год"))
        self.comboBox_7.setItemText(0, _translate("Dialog", "Ничего не выбрано"))
        self.comboBox_7.setItemText(1, _translate("Dialog", "Субъект ГЗ"))
        self.comboBox_7.setItemText(2, _translate("Dialog", "Не субъект ГЗ"))
        self.label_11.setText(_translate("Dialog", "Признак субъекта"))
        self.pushButton.setText(_translate("Dialog", "Запустить парсинг"))


if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    Dialog = QtWidgets.QDialog()
    ui = Ui_Dialog()
    ui.setupUi(Dialog)
    Dialog.show()
    sys.exit(app.exec_())

