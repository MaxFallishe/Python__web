from selenium import webdriver
import time
from openpyxl import load_workbook


excel_row = 1
wb = load_workbook('Empty.xlsx')
ws = wb.active

catalog_name = ""
products_category = ""
group_name = ""


def main():
    global catalog_name

    browser = webdriver.Chrome("D:\\Selenium\\Chrome\\chromedriver.exe")
    browser.maximize_window()

    catalog_link = "https://www.fuchs.com/ru/ru/produkty/katalog-produktov/"
    catalog_list = get_catalog_list(browser, catalog_link)
    catalog_list = catalog_list[:-1]  # Убрать услуги они не нужны

    for catalog_item_link in catalog_list:
        catalog_name = catalog_item_link
        get_categories(browser, catalog_item_link)

    time.sleep(5000)


def get_catalog_list(browser, page_link):
    browser.get(page_link)

    products_catalog = browser.find_element_by_class_name("col-xs-12").find_elements_by_class_name("col-md-3")
    catalog_items_list = []
    for i in products_catalog:
        try:
            catalog_item = i.find_element_by_class_name("csc-textpic-text").find_element_by_tag_name('a').get_attribute("href")
            catalog_items_list.append(catalog_item)
        except Exception as e:
            pass

    print("Catalog successful get")
    return catalog_items_list


def get_categories(browser, link):
    global products_category
    global group_name
    browser.get(link)
    categories = browser.find_element_by_css_selector(".product-filter.product-filter-categories-top").find_elements_by_css_selector(".product-group-teaser.col-md-4.col-sm-6")
    time.sleep(1)

    for i in categories:
        try:
            category = i.find_element_by_css_selector(".product-group-title.teaser-content")
            browser.execute_script("return arguments[0].scrollIntoView(true);", category)
            time.sleep(1)
            category.click()
            print(i.text, "clicked")
            products_category = i.text
            time.sleep(1)
            browser.execute_script("window.scrollTo(0, 0);")
            time.sleep(1)

            # need to get group
            group_button = browser.find_elements_by_css_selector(".btn.dropdown-toggle.btn-default")[1]
            group_button.click()
            print("Button was find and click")

            groups = browser.find_element_by_css_selector(".btn-group.bootstrap-select.form-control.open").find_element_by_css_selector(".dropdown-menu.open").find_elements_by_tag_name("li")
            groups_len = len(groups)

            group_button.click()

            for group in range(0, groups_len):
                group_button = browser.find_elements_by_css_selector(".btn.dropdown-toggle.btn-default")[1]
                group_button.click()

                groups = browser.find_element_by_css_selector(".btn-group.bootstrap-select.form-control.open").find_element_by_css_selector(".dropdown-menu.open").find_elements_by_tag_name("li")
                time.sleep(2)
                group_name = groups[group].text
                groups[group].click()

                get_category_elements(browser)
                time.sleep(5)

        except Exception as e:
            print(e)

    # time.sleep(5000)


def get_category_elements(browser):
    global excel_row
    global products_category
    global catalog_name
    # for first time
    products_table = browser.find_element_by_css_selector(".product-list-wrap.col-xs-12")
    products = browser.find_elements_by_css_selector(".row.product-entry")

    for product in products:
        name = product.find_element_by_css_selector(".col-md-3.col-sm-12").text

        attributes = product.find_element_by_css_selector(".col-md-9.col-sm-12.col-grey")
        information = attributes.find_element_by_class_name("col-xs-12").find_element_by_class_name("product-subtitle").text
        specification = attributes.find_element_by_class_name("product-specifications").text
        permissions = attributes.find_element_by_class_name("product-approvals").text
        recommendations = attributes.find_element_by_class_name("product-recommendations").text
        print(products_category)
        print(group_name)
        print(name)
        #print(information)
        #print(specification)
        #print(permissions)
        #print(recommendations)
        ws.cell(row=excel_row, column=1).value = catalog_name
        ws.cell(row=excel_row, column=2).value = products_category.replace('-', '')
        ws.cell(row=excel_row, column=3).value = group_name
        ws.cell(row=excel_row, column=4).value = name
        ws.cell(row=excel_row, column=5).value = information
        ws.cell(row=excel_row, column=6).value = specification
        ws.cell(row=excel_row, column=7).value = permissions
        ws.cell(row=excel_row, column=8).value = recommendations

        wb.save('Fuchs.xlsx')
        excel_row += 1
    #
    try:
        max_page = int(browser.find_element_by_css_selector(".product-list-pager.product-list-pager-top").find_elements_by_tag_name('li')[-2].text)
        print("MAX Range", max_page)

        for i in range(1, max_page):
            slider_button = browser.find_element_by_css_selector(".product-list-pager.product-list-pager-top").find_elements_by_tag_name('li')[-1].find_element_by_tag_name('a')
            slider_button.click()
            products_table = browser.find_element_by_css_selector(".product-list-wrap.col-xs-12")
            products = browser.find_elements_by_css_selector(".row.product-entry")
            for product in products:
                name = product.find_element_by_css_selector(".col-md-3.col-sm-12").text
                print(name)

            # print(products_table.text)
            print(i)
            time.sleep(5)
    except:
        print("Only one page")


if __name__ == '__main__':
    main()
